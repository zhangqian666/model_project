import openpyxl
import numpy as np
import statsmodels.api as sm
import pandas
import matplotlib.pyplot as plt
from statsmodels.tsa.vector_ar.vecm import VECM

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('/Users/zhangqian/PycharmProjects/model_project/futures_data.xlsx')

ws = wb.active

start_num = 2000
end_num = 2375
# start_num = 2377
# end_num = 2620

first_list = ws['C'][start_num:end_num]

# 初始化 E列表的数据 全部为0
first_list_demo = [0] * (end_num - start_num - 1)

# 将E列表的对象转化为列表
first_list = np.asarray(first_list)

first_demo_index = 0
for first_list_index in range(1, end_num - start_num - 1, 1):
    first_list_demo[first_demo_index] = (first_list[first_list_index].value - first_list[first_list_index - 1].value) / \
                                        first_list[first_list_index - 1].value
    first_demo_index += 1

print(first_list_demo)
# 初始化 第二组数据
second_list = ws['E'][start_num:end_num]

second_list_demo = [0] * (end_num - start_num - 1)

second_list = np.asarray(second_list)

second_demo_index = 0
for second_list_index in range(1, end_num - start_num - 1, 1):
    second_list_demo[second_demo_index] = (second_list[second_list_index].value - second_list[
        second_list_index - 1].value) / second_list[second_list_index - 1].value
    second_demo_index += 1

print(second_list_demo)

# 利用ols模型 构建回归方程

est = sm.OLS(second_list_demo, sm.add_constant(first_list_demo)).fit()



predictList = est.predict()



print(est.predict())

print(est.params)
print(est.summary())

R1 = [0] * (end_num - start_num - 1)

for i in range(0, end_num - start_num - 1):
    R1[i] = second_list_demo[i] - first_list_demo[i] * est.params[1]

var_data = np.var(R1)
second_list_demo = np.asarray(second_list_demo)
var_data_2 = np.var(second_list_demo)

print((var_data_2 - var_data) / var_data_2)

fig = plt.figure()
ax = fig.add_subplot(111)
ax.scatter(first_list_demo, second_list_demo, c='b')
plt.show()
